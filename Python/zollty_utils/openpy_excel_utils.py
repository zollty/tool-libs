import openpyxl
from openpyxl.utils import range_boundaries


def write_excel_with_merged_cells(file_path, data, merged_ranges=None, sheet_name=None):
    """
    将数据写入 Excel 文件，并支持合并单元格。

    :param file_path: str, Excel 文件路径
    :param data: list of lists, 要写入的数据（二维列表）
    :param merged_ranges: list of str, 合并单元格的范围（例如 ['A1:B2', 'C3:D5']），可选
    :param sheet_name: str, 工作表名称，可选，默认为第一个工作表
    """
    # 加载或创建工作簿
    try:
        wb = openpyxl.load_workbook(file_path)
    except FileNotFoundError:
        wb = openpyxl.Workbook()

    # 选择工作表
    if sheet_name:
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.create_sheet(title=sheet_name)
    else:
        ws = wb.active

    # 清空现有内容（避免重复写入）
    for row in ws.iter_rows():
        for cell in row:
            cell.value = None

    # 写入数据
    for row_idx, row_data in enumerate(data, start=1):
        for col_idx, value in enumerate(row_data, start=1):
            # if row_idx == 1269:
            #     print(f"写入数据：行 {row_idx} 列 {col_idx} 值为：{value}")
            ws.cell(row=row_idx, column=col_idx, value=value)

    # 处理合并单元格
    if merged_ranges:
        for merged_range in merged_ranges:
            ws.merge_cells(merged_range)

    # 保存文件
    wb.save(file_path)
    print(f"数据已成功写入文件: {file_path}")


def load_workbook(file_path):
    # 加载工作簿
    try:
        # 正常加载工作簿
        wb = openpyxl.load_workbook(file_path, data_only=True)
    except Exception as e:
        print(f"加载样式失败，尝试跳过样式解析... 错误详情：{e}")
        # 跳过样式解析重新加载
        wb = openpyxl.load_workbook(file_path, data_only=True, keep_links=False)
    # 根据是否指定sheet_name来选择工作表
    # if sheet_name:
    #     return wb[sheet_name]
    # else:
    #     return wb.active
    return wb

def read_sheet_with_merged_cells(ws):
    # 创建一个字典来存储合并单元格的值
    merged_cells_dict = {}

    # 处理所有合并单元格
    for merged_range in ws.merged_cells.ranges:
        min_col, min_row, max_col, max_row = range_boundaries(str(merged_range))
        # 获取合并单元格的值（只取左上角单元格的值）
        merged_value = ws.cell(row=min_row, column=min_col).value

        # 将合并区域内的所有单元格映射到这个值
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                merged_cells_dict[(row, col)] = merged_value

    # 获取数据的行数和列数
    max_row = ws.max_row
    max_col = ws.max_column

    # 存储Excel数据
    excel_data = []

    # 遍历每一行
    for row in range(1, max_row + 1):
        row_data = []

        # 遍历每一列（A、B、C）
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)

            # 检查是否是合并单元格
            if (row, col) in merged_cells_dict:
                value = merged_cells_dict[(row, col)]
            else:
                value = cell.value

            row_data.append(value)

        # 新增：在每行最后添加原始行号（从1开始）
        row_data.append(row)

        # 将行数据添加到结果中
        excel_data.append(row_data)

    # 返回Excel数据
    return excel_data


def __test1__(file_path):
    # 调用函数并获取数据
    ws = load_workbook(file_path)
    data = read_sheet_with_merged_cells(ws.active)

    # 遍历打印数据（最后一列为原始行号）
    for row_index, row_data in enumerate(data, start=1):
        print(f"行 {row_index}（原始行号：{row_data[-1]}）: {row_data}")


def __test2__(file_path):
    # 加载工作簿
    try:
        # 正常加载工作簿
        wb = openpyxl.load_workbook(file_path, data_only=True)
    except Exception as e:
        print(f"加载样式失败，尝试跳过样式解析... 错误详情：{e}")
        # 跳过样式解析重新加载
        wb = openpyxl.load_workbook(file_path, data_only=True, keep_links=False)
    # 获取所有sheet名称
    sheet_names = wb.sheetnames

    # 遍历所有sheet
    for sheet_name in sheet_names:
        ws = wb[sheet_name]
        print(f"\n\n\n处理sheet：{sheet_name}--------------------------------------")
        # 处理每个sheet...
        # 调用函数并获取数据
        data = read_sheet_with_merged_cells(ws)
        # 遍历打印数据（最后一列为原始行号）
        for row_index, row_data in enumerate(data, start=1):
            print(f"行 {row_index}（原始行号：{row_data[-1]}）: {row_data}")


# 使用示例
if __name__ == "__main__":
    # file_path = input("请输入Excel文件路径: ")
    # if not file_path:
    #     file_path = "D:\\__SYNC0-P\\Desktop\\副本-海外营销服系统功能清单及接口.xlsx"
    # # 调用函数并获取数据
    # __test2__(file_path)
    __test1__(r"D:\__SYNC2\git\zollty-misc\tool-libs\Python\vue_dependency_analyzer\api_parser\api_result_20260304_223036.xlsx")
